import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import sys
import datetime
from raw_sheet_functions import format_raw_sheet
from google_functions import share_google_sheet

CURRENT_YEAR = int(datetime.datetime.now().strftime("%Y"))
RETIREMENT_YEAR = 7

RAW_SHEET_NAME, XLSX_FILE, RAW_DF, SHARE_WITH_EMAIL = format_raw_sheet(sys.argv[1])

DATA_SHEET_NAME = "Data"
WORKBOOK = load_workbook(XLSX_FILE)
DATA_SHEET = WORKBOOK.create_sheet(DATA_SHEET_NAME, 0)

COMPANY_COLUMN = "Company"
MACHINE_TYPE_COLUMN = "Type"
MACHINE_STATUS_COLUMN = "Machine Status"
RAM_COLUMN = "RAM"
YEAR_MODEL_COLUMN = "Year Model"

TR_NAME = "Twist Resources"
TT_NAME = "Twist Teams"
TR_DF = RAW_DF[RAW_DF[COMPANY_COLUMN] == TR_NAME]
TT_DF = RAW_DF[RAW_DF[COMPANY_COLUMN] == TT_NAME]

COMPANY = {
    "TR":
        {
            "Name": TR_NAME,
            "DF": TR_DF
        },
    "TT":
        {
            "Name": TT_NAME,
            "DF": TT_DF
        },
    }

COL_START = 1
ROW_START = 1
GAP_SMALL = 2
GAP = 5
GAP_BIG = 7
GAP_HALF = 50

thin = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_machine_count_table(company, starting_row, starting_col):
    machine_type_title_starting_row = starting_row + 2

    DATA_SHEET.cell(row=starting_row, column=starting_col, value=company["Name"]).border = THIN_BORDER

    machine_status_title = ["Active", "Defective", "Spare"]
    raw_machine_status = company["DF"][MACHINE_STATUS_COLUMN].values
    raw_machine_type = company["DF"][MACHINE_TYPE_COLUMN].values
    machine_count = {"Grand Total": {}}

    for i, status in enumerate(raw_machine_status):
        if status not in machine_status_title:
            machine_status_title.append(status)
        machine_type = raw_machine_type[i]
        if machine_type not in machine_count:
            machine_count[machine_type] = {}
        machine_count[machine_type][status] = machine_count[machine_type].get(status, 0) + 1
        machine_count["Grand Total"][status] = machine_count["Grand Total"].get(status, 0) + 1
    machine_status_title.append("Grand Total")

    total_count_data = machine_count.pop("Grand Total")
    machine_count["Grand Total"] = total_count_data

    headers_row_1 = ["Counts of Asset Tag", MACHINE_STATUS_COLUMN]
    headers_row_2 = [MACHINE_TYPE_COLUMN] + machine_status_title

    for i, header in enumerate(headers_row_1):
        DATA_SHEET.cell(row=starting_row + 1, column= starting_col + i, value=header).border = THIN_BORDER
    for i, header in enumerate(headers_row_2):
        DATA_SHEET.cell(row=machine_type_title_starting_row, column=starting_col + i, value=header)

    for i, (machine_type, statuses) in enumerate(machine_count.items()):
        current_row = machine_type_title_starting_row + i + 1
        current_col = starting_col + machine_status_title.index("Grand Total") + 1
        DATA_SHEET.cell(row=current_row, column=starting_col, value=machine_type)
        total_status_count = 0
        for status, count in statuses.items():
            status_index = machine_status_title.index(status) + starting_col + 1
            DATA_SHEET.cell(row=current_row, column=status_index, value=count)
            total_status_count += count
        DATA_SHEET.cell(row=current_row, column=starting_col + machine_status_title.index("Grand Total") + 1, value=total_status_count)

    for row in DATA_SHEET.iter_rows(min_row=machine_type_title_starting_row, min_col=starting_col, max_row=current_row, max_col=current_col):
        for cell in row:
            cell.border = THIN_BORDER

    return(machine_count, current_row, current_col)

def write_machine_summary_table(company, data, title_starting_row, starting_col):
    filtered_data = {"Active": data["Active"], "Spare": data["Spare"], "Sold": data["Sold"]}

    table_starting_row = title_starting_row + 1
    ending_row = table_starting_row + len(filtered_data.items())

    DATA_SHEET.cell(row=title_starting_row, column=starting_col, value=company["Name"] + " Machines Summary").border = THIN_BORDER
    DATA_SHEET.cell(row=table_starting_row, column=starting_col, value="Machine Status").border = THIN_BORDER
    DATA_SHEET.cell(row=table_starting_row, column=starting_col + 1, value="Total").border = THIN_BORDER

    for i, (status, count) in enumerate(filtered_data.items()):
        DATA_SHEET.cell(row=table_starting_row + i + 1, column=starting_col, value=status).border = THIN_BORDER
        DATA_SHEET.cell(row=table_starting_row + i + 1, column=starting_col + 1, value=count).border = THIN_BORDER

    pie_labels = Reference(DATA_SHEET, min_row=table_starting_row + 1, max_row=table_starting_row + len(filtered_data.items()), min_col=starting_col)
    pie_values = Reference(DATA_SHEET, min_row=table_starting_row + 1, max_row=table_starting_row + len(filtered_data.items()), min_col=starting_col + 1)
    pie_title = company["Name"] + " Machines Summary"

    create_pie_chart(pie_title, pie_values, pie_labels, ending_row, starting_col, False)

def write_active_machines_table(company, data, starting_row, starting_col):
    table_starting_row = starting_row + 1
    ending_row = table_starting_row + 1

    DATA_SHEET.cell(row=starting_row, column=starting_col, value=company["Name"] + " Active Machines").border = THIN_BORDER
    DATA_SHEET.cell(row=table_starting_row, column=starting_col, value="Machine Status").border = THIN_BORDER
    DATA_SHEET.cell(row=table_starting_row + 1, column=starting_col, value="Active").border = THIN_BORDER

    active_machine_type_count = 0
    filtered_data = {
        k: v for k, v in data.items()
        if k != "Grand Total" and v.get("Active", 0) >= 1
    }
    for i, (machine_type, statuses) in enumerate(filtered_data.items()):
        DATA_SHEET.cell(row=table_starting_row, column=starting_col + i + 1, value=machine_type).border = THIN_BORDER
        DATA_SHEET.cell(row=table_starting_row + 1, column=starting_col + i + 1, value=statuses["Active"]).border = THIN_BORDER
        active_machine_type_count = i+1

    pie_labels = Reference(DATA_SHEET, min_row=table_starting_row, max_row=table_starting_row, min_col=starting_col + 1, max_col=starting_col + active_machine_type_count)
    pie_values = Reference(DATA_SHEET, min_row=table_starting_row + 1, min_col=starting_col + 1, max_col=starting_col + active_machine_type_count)
    pie_title = company["Name"] + " Active Machines"

    create_pie_chart(pie_title, pie_values, pie_labels, ending_row, starting_col, True)

def write_machine_specs_table(company, starting_row, starting_col, status):
    table_starting_row = starting_row + 1
    DATA_SHEET.cell(row=starting_row, column=starting_col, value=f"{company['Name']} {status} Machines").border = THIN_BORDER
    DATA_SHEET.cell(row=table_starting_row, column=starting_col, value="Year Model").border = THIN_BORDER

    df = company["DF"].copy()
    df[RAM_COLUMN] = df[RAM_COLUMN].astype(str).str.extract(r'(\d+)\s*GB', expand=False).fillna(df[RAM_COLUMN]).astype(str) + "GB"
    filtered_df = df[df[MACHINE_STATUS_COLUMN] == status]
    grouped = filtered_df.groupby([YEAR_MODEL_COLUMN, MACHINE_TYPE_COLUMN, RAM_COLUMN]).size()
    year_index = {}
    machine_type_index = {}
    total_value = {}
    for (year, machine_type, ram), count in grouped.items():
        year = int(year)
        if year not in year_index:
            year_index[year] = len(year_index) + 1
            DATA_SHEET.cell(row=table_starting_row + year_index[year], column=starting_col, value=year).border = THIN_BORDER

        if machine_type not in machine_type_index:
            machine_type_index[machine_type] = len(machine_type_index) + 1
            DATA_SHEET.cell(row=table_starting_row, column=starting_col + machine_type_index[machine_type], value=machine_type).border = THIN_BORDER

        current_row = table_starting_row + year_index[year]
        current_col = starting_col + machine_type_index[machine_type]
        current_value = DATA_SHEET.cell(row=current_row, column=current_col).value
        new_value = f"{current_value}\n{count} - {ram}" if current_value else f"{count} - {ram}"
        DATA_SHEET.cell(row=current_row, column=current_col, value=new_value)
        total_value[machine_type] = total_value.get(machine_type, 0) + count
    ending_col = starting_col + len(machine_type_index) + 1

    DATA_SHEET.cell(row=table_starting_row + len(year_index) + 1, column=starting_col, value="Grand Total").border = THIN_BORDER
    DATA_SHEET.cell(row=table_starting_row, column=ending_col, value="Status").border = THIN_BORDER
    for row in DATA_SHEET.iter_rows(min_row=table_starting_row + 1, min_col=starting_col, max_row=table_starting_row + len(year_index), max_col=ending_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell_year_model = DATA_SHEET.cell(row=cell.row, column=starting_col).value
            if cell.column == ending_col:
                if (CURRENT_YEAR - cell_year_model) >= RETIREMENT_YEAR:
                    DATA_SHEET.cell(row=cell.row, column=ending_col, value="Retirement").border = THIN_BORDER
                elif (CURRENT_YEAR - cell_year_model) == RETIREMENT_YEAR - 1:
                    DATA_SHEET.cell(row=cell.row, column=ending_col, value="Nearing Retirement").border = THIN_BORDER

    for i, (machine_type, count) in enumerate(total_value.items()):
        DATA_SHEET.cell(row=table_starting_row + len(year_index) + 1, column=starting_col + i + 1, value=count).border = THIN_BORDER
    
    return(ending_col)

def create_pie_chart(title, values, labels, starting_row, starting_col, from_rows):
    pie = PieChart()
    pie.dataLabels = DataLabelList()
    chart_row = starting_row + GAP_SMALL
    chart_cell = f"{get_column_letter(starting_col)}{chart_row}"

    pie.add_data(values, from_rows=from_rows)
    pie.set_categories(labels)
    pie.dataLabels.showSerName = False
    pie.dataLabels.showVal = True
    pie.dataLabels.showPercent = False
    pie.dataLabels.showCatName = False
    pie.title = title
    DATA_SHEET.add_chart(pie, chart_cell)

def create_main_data_sheet():
    tr_total_machine_count, tr_machine_count_table_last_row, tr_machine_count_table_last_col = write_machine_count_table(COMPANY["TR"], ROW_START + GAP, COL_START)
    write_machine_summary_table(COMPANY["TR"], tr_total_machine_count["Grand Total"], tr_machine_count_table_last_row + GAP, COL_START)
    write_active_machines_table(COMPANY["TR"], tr_total_machine_count, tr_machine_count_table_last_row + GAP, COL_START + GAP)
    tr_machine_specs_active_last_col = write_machine_specs_table(COMPANY["TR"], ROW_START + GAP_HALF, COL_START, "Active")
    tr_machine_specs_spare_last_col = write_machine_specs_table(COMPANY["TR"], ROW_START + GAP_HALF, tr_machine_specs_active_last_col + GAP_SMALL, "Spare")

    tt_total_machine_count, tt_machine_count_table_last_row, _ = write_machine_count_table(COMPANY["TT"], ROW_START + GAP, tr_machine_count_table_last_col + GAP_SMALL)
    write_machine_summary_table(COMPANY["TT"], tt_total_machine_count["Grand Total"], tt_machine_count_table_last_row + GAP, tr_machine_count_table_last_col + GAP_SMALL)
    write_active_machines_table(COMPANY["TT"], tt_total_machine_count, tt_machine_count_table_last_row + GAP, tr_machine_count_table_last_col + GAP_BIG)
    tt_machine_specs_active_last_col = write_machine_specs_table(COMPANY["TT"], ROW_START + GAP_HALF, tr_machine_specs_spare_last_col + GAP_SMALL, "Active")
    _ = write_machine_specs_table(COMPANY["TT"], ROW_START + GAP_HALF, tt_machine_specs_active_last_col + GAP_SMALL, "Spare")

    dim_holder = DimensionHolder(worksheet=DATA_SHEET)
    for col in range(DATA_SHEET.min_column, DATA_SHEET.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(DATA_SHEET, min=col, max=col, width=35)
    DATA_SHEET.column_dimensions = dim_holder
    WORKBOOK.save(XLSX_FILE)
    share_google_sheet(XLSX_FILE, SHARE_WITH_EMAIL)

create_main_data_sheet()